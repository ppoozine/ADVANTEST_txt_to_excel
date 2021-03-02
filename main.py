import re
import argparse
from openpyxl import Workbook

def get_argments():
    parser = argparse.ArgumentParser()
    parser.add_argument("-o", "--output", dest="xlsx_name", help="Excel Name")
    parser.add_argument("-l", "--load", dest="txt_name", help="Want to parser txt")
    options = parser.parse_args()
    return options

def create_xlsx():
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'X'
    ws['B1'] = 'Y'
    ws['C1'] = 'Category'
    ws['D1'] = 'Parameter'
    ws['E1'] = 'TestID'
    ws['F1'] = 'Value'
    ws['G1'] = 'UP_LIM'
    ws['H1'] = 'LO_LIM'
    ws['I1'] = 'Dpin'
    return wb, ws

def save_xlsx(wb, xlsx_name):
    wb.save(xlsx_name)
    print("[INFO] Successfully established!")

def parser_txt(ws, txt_name):
    data = open(txt_name)
    lines = data.readlines()

    temp_list = []
    find_category = 0
    step = 0
    in_step = 1
    while step != len(lines):
        if re.search('\*+ ADVANTEST DataLog', lines[step]):
            while True:
                if re.search('\s+DUT\s+X\s+Y', lines[step+in_step]):
                    chip_coordinate =  re.search('(\d+)\s+(\d+)\s+(\d+)', lines[step+in_step+1])
                    X = chip_coordinate.group(2)
                    Y = chip_coordinate.group(3)
                    print("X:", X, "\nY:", Y)
                    in_step +=1

                elif lines[step+in_step] == "\n":
                    step = step + in_step
                    in_step = 1
                    break
                else:
                    in_step += 1
        elif re.search('\*+ \[Test', lines[step]):
            fetch_value = True
            while fetch_value == True:
                if re.search('TestID\s+RESULT\s+Value\s+UP_LIM\s+LO_LIM\s+Dpin\s+DUT', lines[step+in_step]):
                    parameter = re.search('"(.*?)"', lines[step]).group(1)
                    print("Parameter:", parameter)
                    find_value_step = 1
                    while True:
                        if lines[step+in_step+find_value_step] == "\n":
                            step = step + in_step + find_value_step
                            fetch_value = False
                            break
                        else:
                            values = re.search('(\S+)\s*(\S+)\s*(\S+)\s*(\S+)\s*(\S+)\s*(\S+)\s*', lines[step+in_step+find_value_step])
                            test_id = values.group(1)
                            temp_dict = {}
                            temp_dict['X'] = X
                            temp_dict['Y'] = Y
                            temp_dict['Parameter'] = parameter
                            temp_dict['TestID'] = values.group(1)
                            temp_dict['Value'] = values.group(2)
                            temp_dict['UP_LIM'] = values.group(3)
                            temp_dict['LO_LIM'] = values.group(4)
                            temp_dict['Dpin'] = values.group(5)
                            temp_list.append(temp_dict)
                            find_value_step += 1

                elif re.search('TestID\s+RESULT\s+Value\s+UP_LIM\s+LO_LIM\s+DUT', lines[step+in_step]):
                    parameter = re.search('"(.*?)"', lines[step]).group(1)
                    print("Parameter:", parameter)
                    find_value_step = 1
                    while True:
                        if lines[step+in_step+find_value_step] == "\n":
                            step = step + in_step + find_value_step
                            fetch_value = False
                            break
                        else:
                            values = re.search('(\S+)\s*(\S+)\s*(\S+)\s*(\S+)\s*(\S+)\s*', lines[step+in_step+find_value_step])
                            test_id = values.group(1)
                            temp_dict = {}
                            temp_dict['X'] = X
                            temp_dict['Y'] = Y
                            temp_dict['Parameter'] = parameter
                            temp_dict['TestID'] = values.group(1)
                            temp_dict['Value'] = values.group(2)
                            temp_dict['UP_LIM'] = values.group(3)
                            temp_dict['LO_LIM'] = values.group(4)
                            temp_dict['Dpin'] = ""
                            temp_list.append(temp_dict)
                            find_value_step += 1

                elif lines[step+in_step] == "\n":
                    step = step + in_step
                    in_step = 1
                    break
                else:
                    in_step += 1
        elif re.search('Category', lines[step]):
            category = re.search('Category.\:.+(\d+)', lines[step]).group(1)
            print('Category:', category)
            for i in range(find_category, len(temp_list)):
                temp_list[i]['Category'] = category
            find_category = i+1
            step +=1
        else:
            step +=1
    
    for i in range(len(temp_list)):
        ws.append([temp_list[i]['X'], temp_list[i]['Y'], temp_list[i]['Category'], temp_list[i]['Parameter'], temp_list[i]['TestID'], temp_list[i]['Value'], temp_list[i]['UP_LIM'], temp_list[i]['LO_LIM'], temp_list[i]['Dpin']])

if __name__ == '__main__':
    options = get_argments()
    wb, ws = create_xlsx()
    parser_txt(ws, options.txt_name)
    save_xlsx(wb, options.xlsx_name)